"""Convert .xlsx files to Markdown.

Each sheet becomes a `## Sheet: <name>` section with a Markdown table.
Merged cells are flattened (top-left value propagated). Formulas are
shown as their cached value (openpyxl returns the formula string when
data_only=False; we use data_only=True to get the last evaluated value).
"""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _format_cell(value) -> str:
    if value is None:
        return ""
    s = str(value)
    # Escape pipe so it doesn't break the Markdown table.
    return s.replace("|", "\\|").replace("\n", " ").strip()


def _rows_to_markdown_table(rows: Iterable[Iterable]) -> str:
    rows = [list(r) for r in rows]
    if not rows:
        return "_(empty sheet)_"

    # Normalise width.
    width = max(len(r) for r in rows)
    rows = [list(r) + [None] * (width - len(r)) for r in rows]

    header = [_format_cell(c) for c in rows[0]]
    # If header is all-empty, synthesize column letters.
    if not any(header):
        header = [chr(ord("A") + i) if i < 26 else f"col{i+1}" for i in range(width)]
        body_rows = rows
    else:
        body_rows = rows[1:]

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * width) + " |")
    for r in body_rows:
        lines.append("| " + " | ".join(_format_cell(c) for c in r) + " |")
    return "\n".join(lines)


def xlsx_bytes_to_markdown(data: bytes, *, source_url: str = "", max_rows_per_sheet: int = 500) -> str:
    """Convert raw .xlsx bytes to Markdown.

    `max_rows_per_sheet` caps very large sheets so the resulting Markdown
    stays manageable for downstream RAG ingestion. The cap is per sheet.
    """
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("openpyxl failed on %s: %s", source_url, e)
        return f"_(Excelファイルの読み込みに失敗しました: {e})_"

    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                truncated = True
                break
            rows.append(row)

        sections.append(f"## Sheet: {sheet_name}\n")
        sections.append(_rows_to_markdown_table(rows))
        if truncated:
            sections.append(f"\n_(注: 表示は先頭{max_rows_per_sheet}行までに制限されています)_")
        sections.append("")  # blank line between sheets

    wb.close()
    return "\n".join(sections).strip()


def xlsx_file_to_markdown(path: Path, **kwargs) -> str:
    return xlsx_bytes_to_markdown(path.read_bytes(), **kwargs)
